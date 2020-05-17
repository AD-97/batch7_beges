import numpy as np
import pandas as pd
import io
from tempfile import NamedTemporaryFile
import openpyxl
import xlrd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook

if __name__ == "__main__":
    import os, sys, inspect

    currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
    parentdir = os.path.dirname(currentdir)
    sys.path.insert(0, parentdir)

from utils.organization_chart import oc
from utils.chorus_dt_handler import ch
from utils.odrive_handler import ov
from utils.osfi_handler import oh


class DataExport:
    def __init__(self, selected_entity):
        self.service = oc.get_entity_by_id(selected_entity)
        self.load_data()

    def load_data(self):
        self.chorus_dt_df = ch.get_structure_data(self.service.code_chorus)
        self.odrive_df = ov.get_structure_data(self.service.code_odrive)
        self.osfi_df = oh.get_structure_data(self.service.code_osfi)
        pass

    def get_file_as_bytes(self):
        """Function returns Excel data as bytes array. It avoids the need to create a file in memory.
            See https://xlsxwriter.readthedocs.io/working_with_pandas.html#additional-pandas-and-excel-information"""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:  # engine="xlsxwriter"
            self.chorus_dt_df.to_excel(writer, sheet_name="data_chorus_dt", index_label="index")
            self.odrive_df.to_excel(writer, sheet_name="data_odrive", index_label="index")
            self.osfi_df.to_excel(writer, sheet_name="data_osfi", index_label="index")
        writer.save()
        xlsx_data = output.getvalue()
        output.seek(0)
        return output

    def get_file_as_bytes_openpyxl(self):
        """Function returns Excel data as bytes array. It avoids the need to create a file in memory.
            https://openpyxl.readthedocs.io/en/stable/tutorial.html"""
        wb = load_workbook("/data/templates/beges_template.xlsx")
        with NamedTemporaryFile() as tmp:
            ws = wb.create_sheet()
            for r in dataframe_to_rows(self.chorus_dt_df, header=True):
                ws.append(r)
            wb.save(tmp.name)
            tmp.seek(0)
            bytes = tmp.read()
        print("INFO: File loaded")
        return io.BytesIO(bytes)


if __name__ == "__main__":
    de = DataExport("5")
    stream = de.get_file_as_bytes_openpyxl()
